from bs4 import BeautifulSoup
from openpyxl import Workbook
import re
import requests
import string
import random
import pandas as pd

wb = Workbook()
ws = wb.active

def random_char(y):
    return ''.join(random.choice(string.ascii_lowercase) for x in range(y))

results=[]
tries = 0
max_tries = 5
ws.append(['NAZWA','KURS','ZMIANA','WYNIK'])
while tries < max_tries:
    shortcut = (random_char(3))
    url = f'https://stooq.pl/q/?s={shortcut}'
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'html.parser')

    table = soup.find('tbody', style='background-color:ffffff')

    if table != None:
        print(str(tries+1)+'/5')

        tries=tries+1    
        rows=table.findAll('tr')
        stock=rows[0].find('td')
        stock=stock.find('span')
        stockResult=''
        if stock:
            stockResult=stock.text

        change=rows[3].find('font', id='c1')
        changeResult=''
        if change:
            changeResult=change.text
        else:
            change=rows[2].findAll('span')
            if change:
                changeResult=change[1].text
            else:
                change=rows[4].findAll('span')
                changeResult=change[1].text

        transaction = rows[8].findAll('span')
        transactionResult=''
        if transaction == []:
            transaction = rows[7].findAll('span')
            if transaction:
                transactionResult=transaction[1].text
            else:
                transaction = rows[6].findAll('span')
                transactionResult=transaction[1].text
        else:
            transactionResult=transaction[1].text
        
        ws.append([shortcut,stockResult, changeResult, transactionResult])

wb.save("gielda.xlsx")


# 2


wb2 = Workbook()
ws2 = wb2.active

url= 'https://www.gry-online.pl/'
page = requests.get(url)
soup = BeautifulSoup(page.text, 'html.parser')
links = soup.find_all('a')
j=1
for link in links:
    try:
        href = link.get('href')
        regex = re.compile('^https?://')
        if regex.match(href):
            title=link.get('title')
            ws2.append([title,href])
            j += 1
    except TypeError:
        pass
    if j >20:
        break


wb2.save("linki.xlsx")

# 3

wb3 = Workbook()
ws3 = wb3.active

url= 'https://www.filmweb.pl/film/Skazani+na+Shawshank-1994-1048'
page = requests.get(url)
soup = BeautifulSoup(page.text, 'html.parser')
director = soup.find('span', itemprop ='name')

boxoffice_section = soup.find('div', class_='filmOtherInfoSection__group')
boxoffice = boxoffice_section.find('div', class_='filmInfo__info')

rating = soup.find('span', class_='filmRating__rateValue')

base = 'https://www.filmweb.pl'
idxStart = len(base)
splitUrl = f'{url[idxStart:]}/dates'
release = soup.find('a', href=splitUrl)
ws3.append([director.text,boxoffice.text,rating.text,release.text])

wb3.save('film.xlsx')